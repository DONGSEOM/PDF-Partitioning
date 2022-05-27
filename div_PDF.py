import os
from PyPDF2 import PdfFileReader, PdfFileWriter
from openpyxl import load_workbook

pdfReader = PdfFileReader('.\sample.pdf', 'rb')
wb = load_workbook('.\sample.xlsx', data_only=True)
ws = wb.active

name_col = ws['A'] #put title raw
names = []

for cell in name_col:
    if cell.value != None:
        names.append(cell.value)

for pageNum in range(pdfReader.numPages):
    pdfWriter = PdfFileWriter()
    page = pdfReader.getPage(pageNum)
    pdfWriter.addPage(page)
    pdfWriter.write(open(f".\{names[pageNum]}.pdf","wb"))
    print('finish')
